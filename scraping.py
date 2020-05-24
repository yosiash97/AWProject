import urllib

from openpyxl import load_workbook
from urllib.request import urlopen
from bs4 import BeautifulSoup
import pandas as pd
from tabulate import tabulate
from dateutil.parser import parse
import requests
import MySQLdb
import re
import json
import mysql.connector
from datetime import datetime
import datetime
from dateutil import parser
import smtplib
from datetime import datetime
from threading import Timer
import datetime
from time import sleep
import schedule
import time
import config
import email.message

# get email to work consistenly
# get rid of colons
# format email message
# more descriptive subject for email

def scrape_eng_pages(filename, sheet, check):
    print("SHEETNAME********:", sheet)
    book = load_workbook(filename)
    old_sheet = sheet
    sheet = book[sheet]
    dictionary = {}
    msg = ""


    # filling up dictionary with city and corresponding website values
    first = False
    for row in sheet.rows:
        dictionary[row[2].value] = row[7].value

    keywords = ['Design', 'Professional ', 'Consult', 'Civil', 'Transportation', 'Bike', 'Pedestrian', 'Sidewalk', 'Street'
                    'Road','Boulevard', 'Blvd', 'Way', 'On-call']

    keywords_track = {}
    pattern = re.compile(
      "(Jan(uary)?|Feb(ruary)?|Mar(ch)?|Apr(il)?|May|Jun(e)?|"
      "Jul(y)?|Aug(ust)?|Sep(tember)?|Oct(ober)?|Nov(ember)?|"
      "Dec(ember)?)\s+\d{1,2},\s+\d{4}")

    regexp = re.compile(pattern)
    dates = []
    #list of websites that either have no current RFPS, or are broken - Can't check the checked/not working field because not all are updated so program will break
    user_agent = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'
    for each in dictionary.keys():
        if each == "City":
            continue
        if each is not None and dictionary[each] is not None and each not in check:
            user_agent = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'

            url = dictionary[each]
            headers = {'User-Agent': user_agent, }

            request = urllib.request.Request(url, None, headers)  # The assembled request
            response = urllib.request.urlopen(request)

            html = response.read()
            soup = BeautifulSoup(html)
            tables = soup.find_all('table')
            final_dates = []
            for table in tables:
                # do your stuff for every table

                try:
                    df = pd.read_html(str(table))
                    if len(df) == 0:
                        continue
                    else:
                        #convert table from website into string paragraphs
                        a = tabulate(df[0], headers='keys', tablefmt='psql')

                        # run through keywords
                        for key in keywords:
                            if key in a:
                                #print("EACH IS IN KEY: ", each, key)
                                if each not in keywords_track:
                                    keywords_track[each] = [key]
                                else:
                                    num_occ = a.count(key)
                                    if not len(keywords_track[each]) == num_occ:
                                        for i in range(num_occ - 1):
                                            keywords_track[each].append(key)

                        if regexp.search(a):
                            print("FOUND DATE!")
                        dates.append((each, re.findall(r"[\d]{1,2}[/.-][\d]{1,2}[/.-][\d]{4}", a), dictionary[each], a))
                except:
                    continue
    print("KEY WORD DICT AFTER FILLING: ", keywords_track)
    array = build_dates(dates)
    print("Array", array)
    email_msg = build_email_msg(array, msg, keywords_track)

    return email_msg

def build_dates(dates):
    final_dates = []
    msg = " "
    for every in dates:
        #print(every)
        for each in every[1]:
            link = every[2]
            body = every[3]

            try:
                datetime_obj = parser.parse(each)

                if datetime_obj > datetime.datetime.today():
                    datetime_obj = datetime_obj.strftime('%m/%d/%Y')
                    final_dates.append((every[0], link, body, datetime_obj))
                    #print("Related Words", every[4])
                    #print("Final Dates Array: ", final_dates)
            except ValueError:
                d = None

    return final_dates
    #build_email_msg(final_dates, msg)


def build_email_msg(final_dates, msg, key_dict):
    latest_rfp = []
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="limited529",
        database="AWDB"
    )
    mycursor = mydb.cursor()
    dict_efficiency = {}

    #cases
    # san francisco: way, plan, road
    #every final date array

    for each in final_dates:
        if each[0] not in dict_efficiency:
            #dict_efficiency[each[0]] = [each[1], each[2], each[3]]
            dict_efficiency[each[0]] = [each[1], each[3]]
        else:
            dict_efficiency[each[0]].append(each[3])

    new_rfp_arr = []
    new_rfp_str = ""
    rfp_keys = []
    for each in dict_efficiency:
        dates_str = ""
        new_rfp_str = ""
        for every in dict_efficiency[each]:
            every = str(every)
            #parsing through dictionary, first value will be website, values from pos 1 to end of array are all dates associated with that link
            if "http" in every and not "https" in every:
                new_rfp_str = every.replace('http://', '')
                new_rfp_arr.append(new_rfp_str)
            elif "https" in every:
                print("BEFORE: ", every)
                new_rfp_str = every.replace('https://', '')
                new_rfp_arr.append(new_rfp_str)
            if "http" not in every and "https" not in every and "www" not in every:
                dates = str(every)
                every = str(every)
                try:
                    keywords = key_dict[each]
                    keywords = ' '.join(keywords)

                    #logic if new listing > prev #get each from db compare length to keywords
                    #mycursor.execute("SELECT * FROM country WHERE Continent = 'Europe'")
                   # mycursor.execute("""SELECT  FROM rfp_watch WHERE link=%s""", ('www.cityofberkeley.info/Finance/Home/Current_Bid_and_Proposal_Opportunities.aspx'))
                    mycursor.execute("""SELECT keycount FROM rfp_watch WHERE link=%s""", (new_rfp_str,))
                    data = mycursor.fetchall()
                    if data is None or data == []:
                        mycursor.execute("""INSERT INTO rfp_watch(link, dates, keycount) VALUES (%s, %s, %s)""",
                                         (new_rfp_str, dates, keywords))
                        rfp_keys.append((new_rfp_str, keywords))
                    else:
                        #string manipulation for keywords returned from dictionary - trying to format to comma separated lists
                        keywords = keywords.replace(',', '')
                        keywords_comma = ' '.join(keywords.split())
                        keywords_comma = keywords_comma.replace(" ", ",")

                        #string manipulation for old keywords that we pulled from DB to prepare to update with new keywords
                        work = data[0]
                        work = str(work[0])
                        commas_added = ', '.join(work.split())
                        old_list = convert_string_to_list(commas_added)
                        keyword_compare = convert_string_to_list(keywords_comma)

                        #if theres a new keyword and that
                        if len(keyword_compare) > len(old_list) and (set(keyword_compare) - set(old_list)):       #new keyword and new rfp

                            difference = (set(keyword_compare) - set(old_list))
                            old_list_str = ' '.join(old_list)
                            difference = ' '.join(difference)
                            mycursor.execute("""UPDATE rfp_watch SET keycount=%s WHERE link=%s""",
                                             (old_list_str + " " + difference, new_rfp_str,))
                            rfp_keys.append((new_rfp_str, difference))        #append a tuple of city, keywords(should be link later)
                        # rfp drop from website but theres a new entry(old db and rfp keywords equal in length)
                        elif len(keyword_compare) == len(old_list) and (set(keyword_compare) - set(old_list)):
                            difference = set(keyword_compare) - set(old_list)
                            difference = ' '.join(difference)
                            old_list_str = ' '.join(old_list)
                            print("There was a potential drop from the RFP website end! ")
                            mycursor.execute("""UPDATE rfp_watch SET keycount=%s WHERE link=%s""",
                                             (difference, new_rfp_str,))
                            rfp_keys.append((new_rfp_str, difference))
                        #if rfp keywords scraped from website are non existant, we must update the corresponding row in the db
                        elif len(keyword_compare) == 0 and len(old_list) != 0:
                            mycursor.execute("DELETE FROM rfp_watch WHERE link=%s", (new_rfp_str,))
                        # if there are additional keywords added to the target website but the sets of new keys & db keys are the same, that means the same key was added to wb
                        elif len(keyword_compare) > len(old_list) and not (set(keyword_compare) - set(old_list)): # new keyword
                            difference = (set(keyword_compare) - set(old_list))
                            old_list_str = ' '.join(old_list)
                            difference = ' '.join(difference)
                            mycursor.execute("""UPDATE rfp_watch SET keycount=%s WHERE link=%s(%s, %s)""",
                                             (old_list_str + " " + difference, new_rfp_str,))
                            rfp_old = str(set(old_list))
                            rfp_keys.append((new_rfp_str, rfp_old))      #since there is no different between the new keyword set and the old set, that means they're all the same return set

                except:
                    print("AHHHHH error inserting")

    mydb.commit()
    latest_rfp = set(latest_rfp)
    latest_rfp = ' '.join(latest_rfp)
    #print("Latest rfp", latest_rfp)
    print("RFP KEYS BEFORE RETURN: ", rfp_keys)
    return  rfp_keys


def email_helper(msg, what, arr):

    #if we're calling send email on rfp scraping
    #string_test = ""
    new_str = ""
    new_arr = []
    for each in arr:
        print("each", each)
        if 'http' in each and not 'https' in each:
            new_str = each.replace('http://', '')
            new_arr.append(new_str)
        elif 'https' in each:
            new_str = each.replace('https://', '')
            new_arr.append(new_str)

    broken_websites_msg = ""
    for each in new_arr:
        broken_websites_msg += each + "\n"

    #if the type of message is of type RFP - parse it
    if what == "RFP":
        msg_tuple = ""
        for each in msg:
            city, keywords = each
            msg_tuple += city + ": "
            for every in keywords:
                msg_tuple += every
            msg_tuple += "\n"

        if msg_tuple:
            try:
                server = smtplib.SMTP('smtp.gmail.com:587')
                server.ehlo()
                server.starttls()
                server.login(config.Email_Username, config.Email_Password)
                message = 'Subject: {}\n\n{}'.format("DAILY RFP REPORT ** NEW RFP LISTINGS **: ", msg_tuple)
                server.sendmail(config.Email_Username, config.To_Username, message)
                server.quit()
                print("Success: Email sent!")
            except:
                print("Email failed to send.")
    # if we are calling send email on the broken websites
    if what == "W":
        if broken_websites_msg:
            try:
                server = smtplib.SMTP('smtp.gmail.com:587')
                server.ehlo()
                server.starttls()
                server.login(config.Email_Username, config.Email_Password)
                message = 'Subject: {}\n\n{}'.format("Daily Broken Website Report: ", broken_websites_msg)
                server.sendmail(config.Email_Username, config.To_Username, message)
                server.quit()
                print("Success: Email sent!")
            except:
                print("Email failed to send.")

def aggregate_broken_websites(filename, sheet):
    book = load_workbook(filename)
    sheet = book[sheet]
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="limited529",
        database="AWDB"
    )
    mycursor = mydb.cursor()

    city = ""
    website = ""
    dictionary = {}
    broken_websites = []
    # filling up dictionary with city and corresponding website values
    first = False
    error_websites_check = []
    for row in sheet.rows:
        dictionary[row[2].value] = row[7].value

    for each in dictionary.keys():
        if each == "City":
            continue
        if each is None or dictionary[each] is None:
            continue
        print(each, dictionary[each])
        user_agent = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'

        try:
            user_agent = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'

            url = dictionary[each]
            headers = {'User-Agent': user_agent, }

            request = urllib.request.Request(url, None, headers)  # The assembled request

            #response = urllib.request.urlopen(request)
            response = urllib.request.urlopen(request, timeout=10).read().decode('utf-8')
        except:
            error_websites_check.append(each)       #appending all broken to an array for the check above
            try:
                mycursor.execute("""INSERT INTO link_watch(city, link) VALUES (%s, %s)""",
                                 (each, dictionary[each]))
                broken_websites.append(dictionary[each])
            except:
                print("The broken website that you are trying to enter is already in the DB")

    mydb.commit()
    mycursor.execute("SELECT city, link FROM link_watch")
    myresult = mycursor.fetchall()
    print("broken websites array", broken_websites)
    mydb.close()
    return error_websites_check,  broken_websites

def dict_formatter(dictionary):
    dict_string = ""
    for k in dictionary:
        dict_string += str(k)
        for each in dictionary[k]:
            dict_string = dict_string + " " + "," + str(each)
        dict_string += "\n"

    return dict_string

def convert_string_to_list(str_input):
    string = ""
    new_list = []
    #print("Len A", len(str_input))
    for i in range(len(str_input)):
        #print(i)
        if str_input[i] != ',':
            string += str_input[i]
        if str_input[i] == ',' or i == len(str_input) - 1:
            #print("in elif")
            #print("i vs len(a)", i, len(str_input))
            new_list.append(string)
            string = ""

    return new_list


def strip_commas_spaces(string):
    if not string:
        return ""
    new_str = ""
    for i in range(len(string)):
        if (string[i] == ' ' or string[i] == ',') and i == len(string) - 2:
            return new_str
        else:
            new_str += string[i]

    return new_str

def job():

    # have the aggregate function return two values
    # check is for all websites from the beginning of time that don't work so that we can skip them when gathering rfps in the scrape_eng_pages func
    # broken is to aggregate the websites that weren't in the DB (check for DB insert - if it doesnt work that means there are duplicates)
    check1, broken1 = aggregate_broken_websites('RLISToriginal.xlsx', 'RFP_V2')
    check2, broken2 = aggregate_broken_websites('RLISToriginal.xlsx', 'RFP_V3')

    #calls rfp web scraper to gather necessary data
    rfp1 = scrape_eng_pages('RLISToriginal.xlsx', 'RFP_V2', check1)
    rfp2 = scrape_eng_pages('RLISToriginal.xlsx', 'RFP_V3', check2)

    rfp_msg = rfp1 + rfp2
    broken_websites_msg = broken1 + broken2

    # #send respective emails (params are weird but I just wanted one send email function so the values of params determines whether to send an
    # #"RFP" subject email or "W" subject emaile
    email_helper(rfp_msg, "RFP", [])
    email_helper("", "W", broken_websites_msg)

job()



#daily scheduler - calls script every morning
schedule.every().day.at("09:01").do(job)

while True:
    schedule.run_pending()
    time.sleep(60)
